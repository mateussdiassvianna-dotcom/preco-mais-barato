# === Imports padrão da biblioteca Python ===
import os
import re
import time
import traceback
import uuid
import unicodedata
from datetime import datetime
from functools import wraps
from decimal import Decimal, InvalidOperation
import requests

# === Bibliotecas externas ===
import pandas as pd
import httpx
from supabase import create_client, Client
from passlib.hash import pbkdf2_sha256
from werkzeug.utils import secure_filename
from flask import (
    Blueprint, render_template, request, redirect, url_for,
    flash, session, jsonify, current_app
)
from flask_login import login_required
from PIL import Image

# === openpyxl para geração de planilhas Excel ===
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Importa db da extensions
from extensions import db

# === BLUEPRINT ===
comerciante_bp = Blueprint("comerciante", __name__, template_folder="../templates")

# === SUPABASE CONFIG ===
SUPABASE_URL = "https://fzlxteusmakgjmurtrur.supabase.co"
SUPABASE_KEY = "SUA_SUPABASE_KEY_AQUI"
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

# === CONFIGURAÇÕES DE UPLOAD ===
UPLOAD_FOLDER = "static/uploads/comerciantes"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
ALLOWED_IMAGE_EXTENSIONS = {"png", "jpg", "jpeg", "gif"}

def allowed_file(filename):
    """Verifica se a extensão da imagem é permitida"""
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_IMAGE_EXTENSIONS

def save_image(file, comerciante_id):
    """Salva imagem simples sem otimização"""
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        folder = os.path.join(UPLOAD_FOLDER, str(comerciante_id))
        os.makedirs(folder, exist_ok=True)
        filepath = os.path.join(folder, filename)
        file.save(filepath)
        return filepath.replace("\\", "/")
    return None

def upload_foto_comerciante(arquivo):
    if not arquivo or arquivo.filename == "":
        return None

    # faz upload via rota unificada
    resp = requests.post(
        url_for("comerciante.upload_imagem_comerciante", _external=True),
        files={"file": arquivo},
        data={"imagem": ""}
    ).json()

    if resp.get("sucesso"):
        return resp.get("url_final")

    return None


def limpar_preco(valor):
    """
    Converte qualquer tipo de preço mal formatado para float.
    Retorna None se o valor não for um número válido.
    
    Compatível com:
    - preços brasileiros (1.234,56)
    - preços internacionais (1,234.56)
    - moedas (R$, $, USD, BRL etc.)
    - números em notação científica (1e3)
    - textos misturados ("5 mil", "R$ 10,0,0")
    - valores com lixo ("Preço?", "null", "NaN")
    """

    if not valor or not str(valor).strip():
        return None

    v = str(valor).strip()

    # Remove tudo que não for número, vírgula, ponto ou sinal
    v = re.sub(r"[^\d,.\-]", "", v)

    # Se tiver vírgula e ponto ao mesmo tempo
    # Decide qual é decimal baseado na última vírgula/ponto
    if "," in v and "." in v:
        if v.rfind(",") > v.rfind("."):
            # Caso BR: 1.234,56
            v = v.replace(".", "").replace(",", ".")
        else:
            # Caso EUA: 1,234.56
            v = v.replace(",", "")

    # Caso padrão BR (12,34)
    elif "," in v:
        v = v.replace(",", ".")

    # Tenta converter para Decimal
    try:
        valor_decimal = Decimal(v)

        # Limite do banco NUMERIC(10,2) → máximo: 99.999.999,99
        if abs(valor_decimal) > Decimal("99999999.99"):
            return None

        # Valor é válido
        return float(valor_decimal)

    except InvalidOperation:
        # Não é número
        return None

def salvar_imagem_otimizada(arquivo, comerciante_id, largura_max=800, qualidade=70):
    if not arquivo or not allowed_file(arquivo.filename):
        return "/static/img/sem-imagem.png"  # fallback automático

    pasta_destino = os.path.join(UPLOAD_FOLDER, str(comerciante_id))
    os.makedirs(pasta_destino, exist_ok=True)

    filename = secure_filename(arquivo.filename)
    caminho_completo = os.path.join(pasta_destino, filename)

    img = Image.open(arquivo)
    if img.width > largura_max:
        proporcao = largura_max / float(img.width)
        altura_nova = int(img.height * proporcao)
        img = img.resize((largura_max, altura_nova), Image.Resampling.LANCZOS)

    img.save(caminho_completo, optimize=True, quality=qualidade)

    return f"/static/uploads/comerciantes/{comerciante_id}/{filename}"

def allowed_import_file(filename):
    """Verifica se o arquivo tem extensão permitida para importação"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ['csv', 'txt', 'xlsx', 'xls']

def normalizar_coluna(nome):
    """
    Normaliza o nome da coluna: minúsculo, sem acentos e substitui espaços por underscores
    """
    nome = str(nome).strip().lower()
    nome = unicodedata.normalize('NFKD', nome).encode('ASCII', 'ignore').decode('utf-8')
    nome = re.sub(r'\s+', '_', nome)
    return nome

def to_float_safe(value):
    if value in (None, "", "null"):
        return None
    try:
        return float(str(value).replace(",", "."))
    except:
        return None


# -----------------------------
# Conexão Supabase
# -----------------------------
SUPABASE_URL = "https://fzlxteusmakgjmurtrur.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImZ6bHh0ZXVzbWFrZ2ptdXJ0cnVyIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc1ODYxMzQzNCwiZXhwIjoyMDc0MTg5NDM0fQ.u8ZzwEFYywapgFWWurYXOb30C87cKLCL_TJ8q1aFjTs"
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)


# -----------------------------
# Decorator de login
# -----------------------------
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            flash("Faça login para acessar essa página.", "erro")
            return redirect(url_for("comerciante.comerciante_login"))
        return f(*args, **kwargs)
    return decorated

# -----------------------------
# Rotas Comerciante
# -----------------------------
@comerciante_bp.route("/opcoes")
def comerciante_opcoes():
    return render_template("comerciante_opcoes.html")


from urllib.parse import quote

from urllib.parse import quote
from flask import render_template, request, redirect, url_for, flash, current_app
from passlib.hash import pbkdf2_sha256
from datetime import datetime

@comerciante_bp.route("/cadastro", methods=["GET", "POST"])
def comerciante_cadastro():
    if request.method == "POST":
        supabase = current_app.config["supabase"]

        print("\n========== INICIANDO CADASTRO ==========\n")

        # Campos obrigatórios
        nome = request.form.get("nome")
        email = request.form.get("email")
        senha = request.form.get("senha")
        print("Nome:", nome)
        print("Email:", email)

        if not nome or not email or not senha:
            print("❌ Erro: Campos obrigatórios faltando")
            flash("Nome, E-mail e Senha são obrigatórios.", "erro")
            return render_template("comerciante_cadastro.html")

        cidade = request.form.get("cidade")
        estado = request.form.get("estado")
        whatsapp = request.form.get("whatsapp")
        endereco_logradouro = request.form.get("endereco_logradouro")
        endereco_numero = request.form.get("endereco_numero")
        endereco_complemento = request.form.get("endereco_complemento")
        faz_entrega = bool(request.form.get("faz_entrega"))
        latitude = to_float_safe(request.form.get("latitude"))
        longitude = to_float_safe(request.form.get("longitude"))

        print("Cidade:", cidade)
        print("Estado:", estado)

        # Verifica duplicado
        print("\n→ Verificando e-mail duplicado...")
        try:
            existing = supabase.table("comerciantes_pendentes").select("id").eq("email", email).execute()
            print("Resultado duplicado:", existing.data)
        except Exception as e:
            print("❌ ERRO AO VERIFICAR DUPLICADO:", e)
            flash("Erro verificando e-mail.", "erro")
            return render_template("comerciante_cadastro.html")

        if existing.data:
            print("❌ E-mail já cadastrado!")
            flash("E-mail já cadastrado e pendente de aprovação.", "erro")
            return render_template("comerciante_cadastro.html")

        # Auth
        print("\n→ Criando usuário no Auth...")
        try:
            auth_user = supabase.auth.sign_up({"email": email, "password": senha})
            user_id = getattr(getattr(auth_user, "user", None), "id", None)
            print("Auth user_id:", user_id)

            if not user_id:
                print("❌ ERRO: Auth retornou user_id vazio!")
                flash("Erro ao gerar ID do Auth.", "erro")
                return render_template("comerciante_cadastro.html")

        except Exception as e:
            print("❌ ERRO NO AUTH:", e)
            flash(f"Erro no Auth: {e}", "erro")
            return render_template("comerciante_cadastro.html")

        # Insert
        novo = {
            "auth_user_id": str(user_id),
            "nome": nome,
            "email": email,
            "senha_hash": pbkdf2_sha256.hash(senha),
            "cidade": cidade,
            "estado": estado,
            "whatsapp": whatsapp,
            "endereco_logradouro": endereco_logradouro,
            "endereco_numero": endereco_numero,
            "endereco_complemento": endereco_complemento,
            "faz_entrega": faz_entrega,
            "latitude": latitude,
            "longitude": longitude,
            "status": "pendente",
            "data_cadastro": datetime.utcnow().isoformat()
        }

        print("\n→ Enviando INSERT para comerciantes_pendentes...")
        print("Payload:", novo)

        try:
            resp = supabase.table("comerciantes_pendentes").insert(novo).execute()
            print("Resposta do insert:", resp)
        except Exception as e:
            print("❌ ERRO NO INSERT:", e)
            flash(f"Erro ao cadastrar: {e}", "erro")
            return render_template("comerciante_cadastro.html")

        # Resposta de erro do supabase
        if getattr(resp, "error", None):
            print("❌ SUPABASE ERROR FIELD:", resp.error)
            flash(f"Erro ao cadastrar: {resp.error}", "erro")
            return render_template("comerciante_cadastro.html")

        print("\n→ Insert realizado com sucesso!")

        # Foto
        if "foto_perfil" in request.files:
            print("\n→ Fazendo upload da foto...")
            try:
                foto_url = upload_foto_comerciante(request.files["foto_perfil"])
                print("Foto URL:", foto_url)
                if foto_url:
                    supabase.table("comerciantes_pendentes").update(
                        {"foto_perfil": foto_url}
                    ).eq("auth_user_id", user_id).execute()
                    print("Foto salva na tabela")
            except Exception as e:
                print("❌ ERRO NO UPLOAD DE FOTO:", e)

        print("\n========== CADASTRO FINALIZADO ==========\n")
        flash("Cadastro realizado com sucesso!", "sucesso")
        return redirect(url_for("comerciante.aprovacao", nome_do_comercio=nome))

    return render_template("comerciante_cadastro.html")


@comerciante_bp.route("/aprovacao")
def aprovacao():
    nome_comercio = request.args.get("nome_do_comercio", "Comércio")
    return render_template("aprovacao.html", nome_comercio=nome_comercio)

from flask import Blueprint, request, session, jsonify, current_app
import os
from werkzeug.utils import secure_filename

@comerciante_bp.route("/login", methods=["GET", "POST"])
def comerciante_login():
    if request.method == "POST":
        supabase = current_app.config["supabase"]
        email = request.form.get("email")
        senha = request.form.get("senha")

        if not email or not senha:
            flash("Por favor, preencha e-mail e senha.", "erro")
            return render_template("comerciante_login.html")

        try:
            # 1️⃣ Login no Supabase Auth
            res = supabase.auth.sign_in_with_password({"email": email, "password": senha})
            user = getattr(res, "user", None) or getattr(getattr(res, "session", None), "user", None)

            if not user or not getattr(user, "id", None):
                flash("E-mail ou senha incorretos. Verifique e tente novamente.", "erro")
                return render_template("comerciante_login.html")

            auth_user_id = str(user.id)

            # 2️⃣ Verifica se e-mail foi confirmado
            if not user.email_confirmed_at:
                flash("Você ainda não confirmou seu e-mail. Verifique sua caixa de entrada e clique no link enviado.", "erro")
                return render_template("comerciante_login.html")

            # 3️⃣ Busca comerciante aprovado usando auth_user_id
            resp = supabase.table("comerciantes").select("*").eq("auth_user_id", auth_user_id).execute()
            dados_usuario = resp.data[0] if resp.data else None

            if not dados_usuario:
                # Verifica se está pendente
                pendente = supabase.table("comerciantes_pendentes").select("id").eq("auth_user_id", auth_user_id).execute()
                if pendente.data:
                    flash("Sua conta ainda está pendente de aprovação pelo administrador. Aguarde a liberação.", "erro")
                else:
                    flash("Usuário não encontrado. Verifique se você se cadastrou corretamente.", "erro")
                return render_template("comerciante_login.html")

            # 4️⃣ Login bem-sucedido
            session["user_id"] = auth_user_id
            session["comerciante_id"] = dados_usuario.get("id")
            flash(f"Olá {dados_usuario.get('nome')}, login realizado com sucesso!", "sucesso")
            return redirect(url_for("comerciante.dashboard_comerciante"))

        except Exception as e:
            flash(f"Ocorreu um erro ao tentar logar: {str(e)}", "erro")
            return render_template("comerciante_login.html")

    return render_template("comerciante_login.html")


from flask import Blueprint, request, jsonify, session, current_app
from werkzeug.utils import secure_filename
import os

@comerciante_bp.route("/editar", methods=["POST"])
def editar_comerciante():
    auth_user_id = session.get("user_id")
    if not auth_user_id:
        return jsonify({"sucesso": False, "erro": "Usuário não está logado"}), 401

    # ------------------------------
    # 👉 Função segura para converter latitude/longitude
    # ------------------------------
    def to_float_safe(value):
        if value is None:
            return None
        value = str(value).strip()
        if value == "" or value.lower() == "null":
            return None
        try:
            return float(value.replace(",", "."))
        except:
            return None

    # ------------------------------
    # 📌 CAMPOS BÁSICOS
    # ------------------------------
    nome = request.form.get("nome")
    cidade = request.form.get("cidade")
    estado = request.form.get("estado")
    whatsapp = request.form.get("whatsapp")
    email_secundario = request.form.get("email_secundario")
    cep = request.form.get("cep")
    descricao = request.form.get("descricao")

    faz_entrega = request.form.get("faz_entrega") in ["1", "true", "True"]

    logradouro = request.form.get("endereco_logradouro")
    numero = request.form.get("endereco_numero")
    complemento = request.form.get("endereco_complemento")

    latitude = request.form.get("latitude")
    longitude = request.form.get("longitude")

    # ------------------------------
    # ⏰ HORÁRIOS
    # ------------------------------
    horario_inicio = request.form.getlist("horario_inicio[]")
    horario_fim = request.form.getlist("horario_fim[]")
    dias_nomes = request.form.getlist("dia_nome[]")
    dias_fechados = request.form.getlist("dia_fechado[]")

    horario_funcionamento = {}
    for i, dia in enumerate(dias_nomes):
        dia = dia.strip()
        if not dia:
            continue

        inicio = horario_inicio[i] if i < len(horario_inicio) else ""
        fim = horario_fim[i] if i < len(horario_fim) else ""
        fechado = dia in dias_fechados

        horario_funcionamento[dia] = {
            "inicio": inicio or "",
            "fim": fim or "",
            "fechado": fechado
        }

    # ------------------------------
    # 🌐 REDES SOCIAIS
    # ------------------------------
    redes_sociais = {
        "site": request.form.get("site") or "",
        "instagram": request.form.get("instagram") or "",
        "facebook": request.form.get("facebook") or "",
        "whatsapp": request.form.get("whatsapp_business") or "",
    }

    # ------------------------------
    # 🖼 FOTO VIA SUPABASE
    # ------------------------------
    foto_perfil = None
    if "foto_perfil" in request.files:
        file = request.files["foto_perfil"]
        if file and file.filename.strip():
            foto_perfil = upload_foto_comerciante(file)

    # ------------------------------
    # 📦 MONTA DADOS
    # ------------------------------
    dados_atualizacao = {
        "nome": nome,
        "cidade": cidade,
        "estado": estado,
        "whatsapp": whatsapp,
        "email_secundario": email_secundario,
        "cep": cep,
        "descricao": descricao,
        "faz_entrega": faz_entrega,
        "endereco_logradouro": logradouro,
        "endereco_numero": numero,
        "endereco_complemento": complemento,
        "latitude": to_float_safe(latitude),
        "longitude": to_float_safe(longitude),
        "horario_funcionamento": horario_funcionamento,
        "redes_sociais": redes_sociais,
    }

    if foto_perfil:
        dados_atualizacao["foto_perfil"] = foto_perfil

    # ------------------------------
    # 💾 ATUALIZA NO SUPABASE
    # ------------------------------
    try:
        supabase = current_app.config["supabase"]

        resp = (
            supabase.table("comerciantes")
            .update(dados_atualizacao)
            .eq("auth_user_id", auth_user_id)
            .execute()
        )

        if not resp.data:
            return jsonify({"sucesso": False, "erro": "Erro ao atualizar perfil"}), 400

        return jsonify({"sucesso": True})

    except Exception as e:
        print("ERRO AO ATUALIZAR:", e)
        return jsonify({"sucesso": False, "erro": str(e)}), 500



@comerciante_bp.route("/logout")
@login_required
def comerciante_logout():
    session.clear()
    flash("Logout realizado com sucesso!", "sucesso")
    return redirect(url_for("comerciante.comerciante_login"))

@comerciante_bp.route("/dashboard")
def dashboard_comerciante():
    if "user_id" not in session:
        flash("Faça login para acessar o dashboard.", "erro")
        return redirect(url_for("comerciante.comerciante_login"))

    supabase = current_app.config["supabase"]
    user_id = session["user_id"]

    try:
        # Busca comerciante aprovado usando auth_user_id
        resp = supabase.table("comerciantes").select("*").eq("auth_user_id", user_id).execute()
        dados_usuario = resp.data[0] if resp.data else None

        if not dados_usuario:
            flash("Erro: usuário não encontrado na base de dados.", "erro")
            return redirect(url_for("comerciante.comerciante_login"))

        # Renderiza o dashboard com os dados do comerciante
        return render_template("dashboard_comerciante.html", usuario=dados_usuario)

    except Exception as e:
        flash(f"Erro ao carregar o dashboard: {e}", "erro")
        return redirect(url_for("comerciante.comerciante_login"))


    try:
        user_id = session["user_id"]

        # Busca o comerciante aprovado na tabela
        resp = supabase.table("comerciantes").select("*").eq("user_id", str(user_id)).execute()
        comerciante = resp.data[0] if resp.data else None

        if not comerciante:
            flash("Comerciante não encontrado ou ainda pendente de aprovação.", "erro")
            return redirect(url_for("comerciante.comerciante_login"))

        # Busca também informações básicas do Supabase Auth (opcional)
        auth_user = None
        try:
            auth_res = supabase.auth.get_user()
            auth_user = getattr(auth_res, "user", None)
        except Exception:
            pass  # Se der erro aqui, apenas ignora — não é crítico

        # Aqui está a correção principal:
        return render_template(
            "dashboard_comerciante.html",
            usuario=comerciante,
            auth_user=auth_user
        )

    except Exception as e:
        flash(f"Erro ao carregar o dashboard: {str(e)}", "erro")
        return redirect(url_for("comerciante.comerciante_login"))



# -----------------------------
# Produtos CRUD
# -----------------------------
@comerciante_bp.route("/produtos")
@login_required
def meus_produtos():
    comerciante_id = session["comerciante_id"]
    resp = supabase.table("produtos").select("*").eq("comerciante_id", comerciante_id).execute()
    produtos = resp.data
    return render_template("meus_produtos.html", produtos=produtos)

@comerciante_bp.route("/api/produtos", methods=["GET"])
@login_required
def api_listar_produtos():
    from flask import jsonify, request, session
    from datetime import datetime, timezone
    import time
    import httpx

    MAX_RETRIES = 3
    RETRY_DELAY = 0.5  # segundos

    try:
        comerciante_id = session.get("comerciante_id")
        if not comerciante_id:
            return jsonify({"sucesso": False, "erro": "Usuário não autenticado"}), 403

        # =========================
        # PARÂMETROS DE BUSCA
        # =========================
        busca = (request.args.get("busca") or "").strip()
        aviso_filtro = request.args.get("aviso", "0") == "1"
        ordenar_por = request.args.get("ordenar_por", "criado_em")
        ordem = request.args.get("ordem", "desc").lower()
        ordem_desc = ordem == "desc"

        # =========================
        # PAGINAÇÃO
        # =========================
        offset = max(int(request.args.get("offset", 0)), 0)
        limite = min(int(request.args.get("limite", 200)), 500)
        end = offset + limite - 1

        total = None

        def executar_query(q):
            for tentativa in range(1, MAX_RETRIES + 1):
                try:
                    return q.execute()
                except (httpx.ReadError, httpx.ConnectError) as e:
                    print(f"⚠️ Tentativa {tentativa} falhou: {e}")
                    time.sleep(RETRY_DELAY * tentativa)
            return None

        # =========================
        # CONTAGEM TOTAL
        # =========================
        if offset == 0:
            count_query = (
                supabase.table("produtos")
                .select("id", count="exact")
                .match({"comerciante_id": comerciante_id})
            )

            if busca:
                count_query = count_query.or_(
                    f"nome.ilike.%{busca}%," f"marca.ilike.%{busca}%," f"categoria.ilike.%{busca}%"
                )

            if aviso_filtro:
                count_query = count_query.or_(
                    "nome.eq.,marca.eq.,categoria.eq.,preco.is.null,imagem.eq."
                )

            count_res = executar_query(count_query)
            total = count_res.count if count_res else 0

        # =========================
        # LISTAGEM DE PRODUTOS
        # =========================
        query = (
            supabase.table("produtos")
            .select("*")
            .match({"comerciante_id": comerciante_id})
            .order(ordenar_por, desc=ordem_desc)
            .range(offset, end)
        )

        if busca:
            query = query.or_(
                f"nome.ilike.%{busca}%," f"marca.ilike.%{busca}%," f"categoria.ilike.%{busca}%"
            )

        if aviso_filtro:
            query = query.or_(
                "nome.eq.,marca.eq.,categoria.eq.,preco.is.null,imagem.eq."
            )

        produtos_res = executar_query(query)
        produtos = produtos_res.data if produtos_res else []

        return jsonify({
            "sucesso": True,
            "offset": offset,
            "limite": limite,
            "total": total,
            "qtd_retorno": len(produtos),
            "produtos": produtos,
            "ultima_atualizacao": datetime.now(timezone.utc).isoformat()
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"sucesso": False, "erro": "Erro ao acessar o Supabase, tente novamente."}), 500

@comerciante_bp.route("/api/produtos/metadados", methods=["GET"])
@login_required
def api_metadados_produtos():
    """
    Retorna as últimas datas de atualização por lote de produtos do comerciante.
    Usado pelo frontend para validar o cache local segmentado.
    """
    from flask import jsonify, request, session
    from datetime import datetime, timezone

    try:
        comerciante_id = session.get("comerciante_id")
        if not comerciante_id:
            return jsonify({"sucesso": False, "erro": "Usuário não autenticado"}), 403

        limite = min(int(request.args.get("limite", 500)), 2000)
        ordenar_por = request.args.get("ordenar_por", "criado_em")
        ordem = request.args.get("ordem", "desc").lower()
        desc_flag = ordem == "desc"

        # Busca todos os IDs e timestamps básicos, sem sobrecarregar a resposta
        resp = (
            supabase.table("produtos")
            .select("id, atualizado_em, criado_em")
            .eq("comerciante_id", comerciante_id)
            .order(ordenar_por, desc=desc_flag)
            .execute()
        )

        produtos = resp.data or []
        if not produtos:
            return jsonify({"sucesso": True, "lotes": {}})

        # Divide em lotes e calcula o timestamp mais recente de cada lote
        lotes = {}
        for i in range(0, len(produtos), limite):
            lote_num = (i // limite) + 1
            sub = produtos[i:i + limite]

            datas = [
                p.get("atualizado_em") or p.get("criado_em")
                for p in sub
                if p.get("atualizado_em") or p.get("criado_em")
            ]
            if not datas:
                continue

            try:
                mais_recente = max(
                    datetime.fromisoformat(d).astimezone(timezone.utc)
                    if isinstance(d, str) else d
                    for d in datas
                )
                lotes[str(lote_num)] = mais_recente.isoformat()
            except Exception:
                lotes[str(lote_num)] = datetime.now(timezone.utc).isoformat()

        return jsonify({"sucesso": True, "lotes": lotes})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            "sucesso": False,
            "erro": str(e)
        }), 500



@comerciante_bp.route("/api/produtos", methods=["POST"])
@login_required
def api_criar_produto():
    try:
        dados = request.get_json() if request.is_json else request.form
        nome = (dados.get("nome") or "").strip()
        preco_raw = dados.get("preco")

        if not nome or not preco_raw:
            return jsonify({"sucesso": False, "erro": "Nome e preço são obrigatórios."}), 400

        try:
            preco = float(str(preco_raw).replace(",", "."))
        except Exception:
            return jsonify({"sucesso": False, "erro": "Preço inválido."}), 400

        # Upload da imagem
        imagem = None
        if "imagem" in request.files:
            arquivo = request.files["imagem"]
            if arquivo.filename:
                filename = secure_filename(arquivo.filename)
                pasta_upload = os.path.join("static", "uploads")
                os.makedirs(pasta_upload, exist_ok=True)
                caminho_arquivo = os.path.join(pasta_upload, filename)
                arquivo.save(caminho_arquivo)
                imagem = f"uploads/{filename}"

        produto = {
            "comerciante_id": session["comerciante_id"],
            "nome": nome,
            "preco": preco,
            "marca": (dados.get("marca") or "").strip(),
            "categoria": (dados.get("categoria") or "").strip(),
            "descricao": (dados.get("descricao") or "").strip(),
            "imagem": imagem or (dados.get("imagem") or "").strip(),
            "criado_em": datetime.utcnow().isoformat(),
            "atualizado_em": datetime.utcnow().isoformat()
        }

        resp = supabase.table("produtos").insert(produto).execute()

        if hasattr(resp, "error") and resp.error:
            return jsonify({"sucesso": False, "erro": resp.error.message}), 500

        if not resp.data:
            return jsonify({"sucesso": False, "erro": "Erro ao salvar no banco."}), 500

        return jsonify({"sucesso": True, "produto": resp.data[0]}), 201

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"sucesso": False, "erro": str(e)}), 500


# ---- Rota de edição de produto ----
@comerciante_bp.route("/api/produtos/<id>", methods=["PUT"])
@login_required
def api_editar_produto(id):
    try:
        dados = request.get_json() if request.is_json else request.form
        updates = {}

        for campo in ["nome", "preco", "marca", "categoria", "descricao"]:
            if campo in dados:
                updates[campo] = dados[campo].strip() if isinstance(dados[campo], str) else dados[campo]

        if "preco" in updates:
            try:
                updates["preco"] = float(str(updates["preco"]).replace(",", "."))
            except:
                return jsonify({"sucesso": False, "erro": "Preço inválido"}), 400

        # Upload da imagem ou fallback
        if "imagem" in request.files and request.files["imagem"].filename:
            arquivo = request.files["imagem"]
            updates["imagem"] = salvar_imagem_otimizada(arquivo, session["comerciante_id"])
        elif "imagem" in dados and dados["imagem"].strip():
            updates["imagem"] = dados["imagem"].strip()
        else:
            updates["imagem"] = "/static/img/sem-imagem.png"  # fallback

        updates["atualizado_em"] = datetime.utcnow().isoformat()

        resp = supabase.table("produtos") \
            .update(updates) \
            .eq("id", str(id)) \
            .eq("comerciante_id", session["comerciante_id"]) \
            .execute()

        if not resp.data:
            return jsonify({"sucesso": False, "erro": "Produto não encontrado ou não pertence ao comerciante."}), 404

        return jsonify({"sucesso": True, "produto": resp.data[0]}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"sucesso": False, "erro": str(e)}), 500

    
@comerciante_bp.route("/api/produtos/<id>", methods=["DELETE"])
@login_required
def api_deletar_produto(id):
    # 1️⃣ Busca o produto para pegar o nome da imagem
    produto = supabase.table("produtos") \
        .select("id, imagem") \
        .eq("id", id) \
        .eq("comerciante_id", session["comerciante_id"]) \
        .single() \
        .execute()

    if not produto.data:
        return jsonify({"erro": "Produto não encontrado"}), 404

    # 2️⃣ Deleta a imagem do bucket se existir
    if produto.data.get("imagem"):
        nome_arquivo = produto.data["imagem"].split("/")[-1]  # pega só o nome do arquivo
        supabase.storage.from_("produtos").remove([nome_arquivo])

    # 3️⃣ Deleta o produto do banco
    resp = supabase.table("produtos") \
        .delete() \
        .eq("id", id) \
        .eq("comerciante_id", session["comerciante_id"]) \
        .execute()

    return jsonify({"sucesso": True})

# -----------------------------
# Importação de planilha
# -----------------------------
@comerciante_bp.route("/api/produtos/importar", methods=["POST"])
@login_required
def importar_produtos():
    import pandas as pd
    import chardet, traceback, os, unicodedata, xml.etree.ElementTree as ET, numpy as np
    from werkzeug.utils import secure_filename
    from datetime import datetime
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import jsonify, request, session, send_file

    caminho_temp = None
    wb_raw = None

    try:
        # ===================== Arquivo enviado =====================
        if "arquivo" not in request.files:
            return jsonify({"sucesso": False, "erro": "Nenhum arquivo enviado."}), 400

        arquivo = request.files["arquivo"]
        if arquivo.filename == "":
            return jsonify({"sucesso": False, "erro": "Nenhum arquivo selecionado."}), 400

        filename = secure_filename(arquivo.filename)
        extensao = filename.lower().split(".")[-1]
        caminho_temp = os.path.join("static/uploads/temp", filename)
        os.makedirs(os.path.dirname(caminho_temp), exist_ok=True)
        arquivo.save(caminho_temp)

        # ===================== Leitura do arquivo =====================
        rows_raw = []
        df = None

        # -----------------------------------------------------------------
        # ===================== XLS / XLSX ===============================
        # -----------------------------------------------------------------
        if extensao in ["xlsx", "xls"]:

            # --- Se for XLS antigo, converte automaticamente ---
            if extensao == "xls":
                try:
                    df_xls = pd.read_excel(caminho_temp, dtype=str, engine="xlrd")
                    caminho_convertido = caminho_temp + ".xlsx"
                    df_xls.to_excel(caminho_convertido, index=False)
                    caminho_temp = caminho_convertido
                    extensao = "xlsx"
                except Exception as e:
                    return jsonify({
                        "sucesso": False,
                        "erro": f"Falha ao ler arquivo .xls. Converta para .xlsx. Detalhes: {str(e)}"
                    }), 400

            # --- Agora lê normalmente como XLSX ---
            wb_raw = load_workbook(caminho_temp, read_only=True, data_only=True)
            ws_raw = wb_raw[wb_raw.sheetnames[0]]

            header = []
            for row in ws_raw.iter_rows(min_row=1, max_row=1, values_only=True):
                header = [str(c).strip() if c is not None else "" for c in row]

            for row in ws_raw.iter_rows(min_row=2, values_only=True):
                r = {}
                for i, val in enumerate(row):
                    col_name = header[i] if i < len(header) else f"col{i}"
                    if val is None:
                        r[col_name.strip().lower()] = ""
                    else:
                        if isinstance(val, int):
                            r[col_name.strip().lower()] = str(val)
                        elif isinstance(val, float):
                            r[col_name.strip().lower()] = str(int(val)) if val.is_integer() else str(val)
                        else:
                            r[col_name.strip().lower()] = str(val)
                rows_raw.append(r)

            try:
                wb_raw.close()
            except Exception:
                pass
            wb_raw = None

            df = pd.DataFrame(rows_raw) if rows_raw else pd.DataFrame()

        # -----------------------------------------------------------------
        # ===================== CSV ===============================
        # -----------------------------------------------------------------
        elif extensao == "csv":
            with open(caminho_temp, "rb") as f:
                result = chardet.detect(f.read(10000))
            encoding = result["encoding"] or "utf-8"

            tentativas = [";", ",", "\t", "|"]
            for sep in tentativas:
                try:
                    df = pd.read_csv(caminho_temp, encoding=encoding, sep=sep, dtype=str)
                    if len(df.columns) > 0:
                        break
                except Exception:
                    df = None
                    continue

            if df is None:
                return jsonify({"sucesso": False, "erro": "Não foi possível parsear o CSV."}), 400

            df = df.fillna("")
            rows_raw = df.astype(str).to_dict(orient="records")

        # -----------------------------------------------------------------
        # ===================== JSON ===============================
        # -----------------------------------------------------------------
        elif extensao == "json":
            df = pd.read_json(caminho_temp, dtype=False)
            df = df.fillna("")
            rows_raw = df.astype(str).to_dict(orient="records")

        # -----------------------------------------------------------------
        # ===================== XML ===============================
        # -----------------------------------------------------------------
        elif extensao == "xml":
            tree = ET.parse(caminho_temp)
            root = tree.getroot()
            data = [{child.tag: child.text or "" for child in item} for item in root]
            df = pd.DataFrame(data).fillna("")
            rows_raw = df.astype(str).to_dict(orient="records")

        # -----------------------------------------------------------------
        # ===================== TXT ===============================
        # -----------------------------------------------------------------
        elif extensao == "txt":
            with open(caminho_temp, "rb") as f:
                result = chardet.detect(f.read(10000))

            encoding = result["encoding"] or "utf-8"

            with open(caminho_temp, "r", encoding=encoding, errors="ignore") as f:
                primeira_linha = f.readline()

            if ";" in primeira_linha or "," in primeira_linha or "\t" in primeira_linha:
                tentativas = [";", ",", "\t", "|"]
                parsed = False
                for sep in tentativas:
                    try:
                        df = pd.read_csv(caminho_temp, encoding=encoding, sep=sep, dtype=str)
                        if len(df.columns) > 0:
                            parsed = True
                            break
                    except Exception:
                        continue

                if not parsed:
                    linhas = [l.strip() for l in open(caminho_temp, encoding=encoding, errors="ignore").readlines() if l.strip()]
                    rows_raw = [{"linha": l} for l in linhas]
                    df = pd.DataFrame(rows_raw)
                else:
                    df = df.fillna("")
                    rows_raw = df.astype(str).to_dict(orient="records")

            else:
                linhas = [l.strip() for l in open(caminho_temp, encoding=encoding, errors="ignore").readlines() if l.strip()]
                rows_raw = [{"linha": l} for l in linhas]
                df = pd.DataFrame(rows_raw)

        # -----------------------------------------------------------------
        # ===================== ODS ===============================
        # -----------------------------------------------------------------
        elif extensao == "ods":
            df = pd.read_excel(caminho_temp, engine="odf", dtype=str).fillna("")
            rows_raw = df.astype(str).to_dict(orient="records")

        else:
            return jsonify({"sucesso": False, "erro": f"Formato de arquivo não suportado: .{extensao}"}), 400

        # -----------------------------------------------------------------
        # ===================== Arquivo vazio ===============================
        # -----------------------------------------------------------------
        if df is None or df.empty:
            return jsonify({"sucesso": False, "erro": "Arquivo vazio ou não foi possível ler."}),

        # ===================== Normalização =====================
        def normalizar_coluna(c):
            c = str(c).strip().lower()
            c = unicodedata.normalize("NFKD", c).encode("ASCII", "ignore").decode("utf-8")
            return c

        def normalizar_texto(valor):
            if valor is None:
                return ""
            v = str(valor).strip().lower()
            v = unicodedata.normalize("NFKD", v).encode("ASCII", "ignore").decode("utf-8")
            return v

        df.columns = [normalizar_coluna(c) for c in df.columns]

        rows_raw_norm = []
        for r in rows_raw:
            rn = {}
            for k, v in r.items():
                rn[normalizar_coluna(k)] = v if v is not None else ""
            rows_raw_norm.append(rn)
        rows_raw = rows_raw_norm

        df = df.replace({np.nan: None})

        comerciante_id = session.get("comerciante_id")
        imagem_padrao = "/static/img/sem-imagem.png"
        produtos_importados, erros = [], []

        # ===================== Produtos já existentes =====================
        nomes_existentes_resp = supabase.table("produtos").select("nome").eq("comerciante_id", comerciante_id).execute()
        nomes_existentes = {p["nome"].strip().lower() for p in (nomes_existentes_resp.data or [])}

        def traduzir_erro(e: Exception):
            erro_str = str(e)
            if "could not convert string to float" in erro_str or "invalid literal for float" in erro_str:
                return "O campo de preço contém texto ou está vazio. Verifique a coluna Preço."
            elif "NoneType" in erro_str:
                return "Um dos campos obrigatórios (como nome ou preço) está vazio."
            return f"Erro inesperado: {erro_str}"

        def processar_imagem(valor):
            if not valor or str(valor).strip().lower() in ["nan", "none", ""]:
                return imagem_padrao
            valor = str(valor).strip()
            if valor.startswith("http://") or valor.startswith("https://"):
                return valor
            local = os.path.join("static/uploads", valor)
            if os.path.exists(local):
                return "/" + local.replace(os.sep, "/")
            return imagem_padrao

        vistos_no_arquivo = {}
        primeira_ocorrencia_info = {}
        duplicados_internos = []
        nome_sem_marca_linhas = {}

        # ===================== Processamento =====================
        for idx, row in df.iterrows():
            linha = idx + 2
            try:
                nome_raw = row.get("nome", "")
                nome = str(nome_raw).strip() if nome_raw is not None else ""
                marca_raw = row.get("marca", "") or ""
                marca = str(marca_raw).strip()
                categoria_raw = row.get("categoria", "") or ""
                categoria = str(categoria_raw).strip()
                descricao_raw = row.get("descricao", "") or ""
                descricao = str(descricao_raw).strip()

                preco_texto = ""
                if idx < len(rows_raw) and "preco" in rows_raw[idx]:
                    preco_texto = rows_raw[idx].get("preco", "")
                else:
                    preco_val = row.get("preco", "")
                    preco_texto = "" if preco_val is None else str(preco_val)

                preco_texto = str(preco_texto).strip()

                preco_num = None
                try:
                    tmp = preco_texto.replace("R$", "").replace("r$", "").strip()
                    tmp = tmp.replace("\u00A0", "").replace(" ", "")
                    if tmp == "" or tmp.lower() in ["nan", "none"]:
                        preco_num = None
                    else:
                        if "." in tmp and "," in tmp:
                            tmp = tmp.replace(".", "").replace(",", ".")
                        elif "," in tmp and tmp.count(",") == 1 and "." not in tmp:
                            tmp = tmp.replace(",", ".")
                        preco_num = float(tmp)
                except Exception:
                    preco_num = None

                if not nome:
                    erros.append({"linha": linha, "erro": "Nome ausente. Produto sem nome.", "tipo": "erro"})
                    continue

                nome_norm = normalizar_texto(nome)
                marca_norm = normalizar_texto(marca)
                chave = (nome_norm, marca_norm)

                if marca_norm == "":
                    if nome_norm not in nome_sem_marca_linhas:
                        nome_sem_marca_linhas[nome_norm] = []
                    nome_sem_marca_linhas[nome_norm].append(linha)

                if chave in vistos_no_arquivo:
                    primeira_linha = vistos_no_arquivo[chave]
                    erros.append({
                        "linha": linha,
                        "erro": f"Produto duplicado na própria planilha. Primeira ocorrência na linha {primeira_linha}.",
                        "tipo": "erro"
                    })
                    duplicados_internos.append({
                        "linha_duplicada": linha,
                        "nome": nome,
                        "marca": marca or "(vazio)",
                        "primeira_ocorrencia": primeira_linha
                    })
                    continue
                else:
                    vistos_no_arquivo[chave] = linha
                    primeira_ocorrencia_info[chave] = {
                        "linha": linha,
                        "nome": nome,
                        "marca": marca or "",
                        "categoria": categoria or "",
                        "preco_texto": preco_texto,
                        "preco": preco_num if preco_num is not None else "",
                        "descricao": descricao or ""
                    }

                if (preco_texto == "" or preco_texto.lower() in ["nan", "none"]) and preco_num is None:
                    erros.append({"linha": linha, "erro": "Preço ausente ou inválido.", "tipo": "erro"})
                    continue

                produtos_importados.append({
                    "comerciante_id": comerciante_id,
                    "nome": nome,
                    "preco": preco_num if preco_num is not None else None,
                    "marca": marca,
                    "categoria": categoria,
                    "descricao": descricao,
                    "imagem": processar_imagem(row.get("imagem")),
                    "criado_em": datetime.utcnow().isoformat(),
                    "atualizado_em": datetime.utcnow().isoformat(),
                })
                nomes_existentes.add(nome.lower())

            except Exception as e:
                erros.append({"linha": linha, "erro": traduzir_erro(e), "tipo": "erro"})

        # ===================== Pós-processamento duplicados sem marca =====================
        nome_to_chaves = {}
        for (nome_norm, marca_norm), primeira_linha in vistos_no_arquivo.items():
            if nome_norm not in nome_to_chaves:
                nome_to_chaves[nome_norm] = []
            nome_to_chaves[nome_norm].append((marca_norm, primeira_linha))

        for nome_norm, lista_chaves in nome_to_chaves.items():
            if len(lista_chaves) > 1:
                marcas_vazias = [t for t in lista_chaves if t[0] == ""]
                if marcas_vazias:
                    linhas_sem_marca = nome_sem_marca_linhas.get(nome_norm, [])
                    for linha_sem_marca in linhas_sem_marca:
                        ja_tem = any(e.get("linha") == linha_sem_marca and "marca" in e.get("erro", "").lower() for e in erros)
                        if not ja_tem:
                            erros.append({
                                "linha": linha_sem_marca,
                                "erro": "Nome presente em outras linhas, mas MARCA ausente aqui — impossível decidir se é o mesmo produto.",
                                "tipo": "erro"
                            })
                            duplicados_internos.append({
                                "linha_duplicada": linha_sem_marca,
                                "nome": primeira_ocorrencia_info.get((nome_norm, ""), {}).get("nome", ""),
                                "marca": "(vazio)",
                                "primeira_ocorrencia": primeira_ocorrencia_info.get((nome_norm, lista_chaves[0][0]), {}).get("linha", "")
                            })

        # ===================== Inserção no banco =====================
        if produtos_importados:
            supabase.table("produtos").insert(produtos_importados).execute()

        # ===================== Relatório (Excel) =====================
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Importação"

        cor_titulo = "2F4865"
        cor_secao = "D9E1F2"
        cor_texto = "1F1F1F"
        cor_erro = "FFDDDD"
        cor_erro_header = "C00000"
        cor_aviso = "FFF6CC"
        cor_sucesso = "E8F6EF"

        fonte_titulo = Font(bold=True, size=16, color="FFFFFF")
        fonte_header = Font(bold=True, color=cor_texto)
        fonte_normal = Font(size=11, color=cor_texto)
        fonte_secao = Font(bold=True, size=12, color="000000")
        fonte_dica = Font(italic=True, size=11, color="000000")
        alin_esquerda = Alignment(horizontal="left", vertical="top", wrap_text=True)
        alin_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
        borda = Border(
            left=Side(style="thin", color="999999"),
            right=Side(style="thin", color="999999"),
            top=Side(style="thin", color="999999"),
            bottom=Side(style="thin", color="999999")
        )

        # limpa sheet
        ws.delete_rows(1, ws.max_row)
        ws.merge_cells("A1:G1")
        ws["A1"].value = "Relatório Analítico de Validação de Produtos - 🔁 Preço Mais Barato"
        ws["A1"].font = fonte_titulo
        ws["A1"].alignment = alin_centro
        ws["A1"].fill = PatternFill(start_color=cor_titulo, end_color=cor_titulo, fill_type="solid")
        ws.row_dimensions[1].height = 26

        # === CABEÇALHO VERTICAL ===
        cab = {
            "Data do processamento": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            "Tipo de operação": "IMPORTAÇÃO DE PRODUTOS",
            "Sistema": "Preço Mais Barato - Núcleo de Inteligência",
            "Versão do motor": "2.3.1",
            "Relatório": "Relatório Analítico de Validação de Produtos"
        }
        for k, v in cab.items():
            ws.append([k, v])
            for cell in ws[ws.max_row]:
                cell.font = fonte_normal
                cell.alignment = alin_esquerda
        ws.append([])

        # ===== Resumo =====
        ws.append(["Resumo da Importação"])
        ws["A" + str(ws.max_row)].font = fonte_secao
        ws["A" + str(ws.max_row)].fill = PatternFill(start_color=cor_secao, end_color=cor_secao, fill_type="solid")

        ws.append(["Total de linhas", "Importados", "Com Erro", "Com Aviso"])
        for c in ws[ws.max_row]:
            c.font = fonte_header
            c.alignment = alin_centro
            c.border = borda

        ws.append([
            len(df),
            len(produtos_importados),
            len([e for e in erros if e.get("tipo") == "erro"]),
            len([e for e in erros if e.get("tipo") == "aviso"])
        ])
        for c in ws[ws.max_row]:
            c.alignment = alin_centro
            c.border = borda

        ws.append([])

        # ===== Produtos com Erro / Avisos =====
        ws.append(["⚠️ Produtos com Erro / Avisos (Simples ➜ Solução)"])
        ws["A" + str(ws.max_row)].font = fonte_secao
        ws["A" + str(ws.max_row)].fill = PatternFill(start_color=cor_secao, end_color=cor_secao, fill_type="solid")

        if erros:
            ws.append(["Linha", "O que aconteceu (simples)", "O que fazer (simples)", "Análise técnica (código)"])
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = alin_centro
                cell.fill = PatternFill(start_color=cor_erro_header, end_color=cor_erro_header, fill_type="solid")
                cell.border = borda

            # agrupar erros por linha
            erros_por_linha = {}
            for e in erros:
                l = e.get("linha")
                msg = e.get("erro", "")
                if l not in erros_por_linha:
                    erros_por_linha[l] = []
                if msg not in erros_por_linha[l]:
                    erros_por_linha[l].append(msg)

            def mapear_instrucoes(joined_msg):
                lower_msg = joined_msg.lower()
                instrucoes = []
                instr_tecnica = []

                if "nome ausente" in lower_msg or "sem nome" in lower_msg:
                    instrucoes.append("Escreva o NOME do produto nessa linha.")
                    instr_tecnica.append(("E-01", "Nome ausente: campo obrigatório."))

                if "preço ausente" in lower_msg or "preço inválido" in lower_msg:
                    instrucoes.append("Preencha o PREÇO (ex.: 2,99 ou 299).")
                    instr_tecnica.append(("E-03", "Preço ausente/inválido: impede cadastro."))

                if "duplicado na própria planilha" in lower_msg:
                    instrucoes.append("Remova duplicatas ou mantenha apenas a primeira ocorrência.")
                    instr_tecnica.append(("E-04", "Duplicata exata (nome+marca)."))

                if "já existe" in lower_msg:
                    instrucoes.append("Produto já cadastrado. Use 'Atualizar Produtos' para alterar.")
                    instr_tecnica.append(("E-05", "Conflito com base: nome já presente."))

                if "marca ausente" in lower_msg:
                    instrucoes.append("Preencha MARCA; se não souber escreva 'SEM MARCA'.")
                    instr_tecnica.append(("E-06", "Marca ausente: dificulta distinção."))

                if not instrucoes:
                    instrucoes.append("Corrija os campos indicados e importe novamente.")
                    instr_tecnica.append(("E-00", "Erro genérico: verificar manualmente."))

                return " ".join(instrucoes), " | ".join([f"{c}: {d}" for c, d in instr_tecnica])

            for linha_idx in sorted(erros_por_linha.keys()):
                mensagens = erros_por_linha[linha_idx]
                joined_msg = "; ".join(mensagens)
                instr_simples, instr_tecnica = mapear_instrucoes(joined_msg)

                ws.append([linha_idx, joined_msg, instr_simples, instr_tecnica])
                last_row = ws.max_row

                fill = PatternFill(start_color=cor_erro, end_color=cor_erro, fill_type="solid")
                if "marca ausente" in joined_msg.lower():
                    fill = PatternFill(start_color=cor_aviso, end_color=cor_aviso, fill_type="solid")

                for cell in ws[last_row]:
                    cell.alignment = alin_esquerda
                    cell.font = fonte_normal
                    cell.border = borda
                    cell.fill = fill

            ws.append([])
            ws.append(["🔎 Resumo de Duplicados - Explicação Simples"])
            ws["A" + str(ws.max_row)].font = fonte_secao
            ws.append(["Aqui listamos os duplicados detectados e uma explicação clara do que fazer."])

            if duplicados_internos:
                dup_by_name = {}
                for d in duplicados_internos:
                    nome = d.get("nome", "")
                    dup_by_name.setdefault(nome, []).append(d)

                for nome, items in dup_by_name.items():
                    linhas = sorted(set([
                        it.get("linha_duplicada") for it in items
                    ] + [
                        it.get("primeira_ocorrencia") for it in items if it.get("primeira_ocorrencia")
                    ]))
                    linhas = [str(x) for x in linhas if x]

                    ws.append([f"Produto: {nome} → Linhas envolvidas: {', '.join(linhas)}"])
                    ws.append(["  O que significa: Esse nome apareceu mais de uma vez. Verifique marcas e preços."])
                    ws.append(["  Como resolver: mantenha a linha correta e corrija/remova as demais."])
            else:
                ws.append(["Nenhum duplicado interno detalhado."])

        else:
            ws.append(["✅ Nenhum erro encontrado."])
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True, color="000000")
                cell.alignment = alin_centro

        ws.append([])

        # ===== Produtos importados =====
        ws.append(["✅ Produtos Importados com Sucesso"])
        ws["A" + str(ws.max_row)].font = fonte_secao
        ws["A" + str(ws.max_row)].fill = PatternFill(start_color=cor_sucesso, end_color=cor_sucesso, fill_type="solid")

        ws.append(["Nome", "Preço (texto original)", "Preço (valor)", "Categoria", "Marca", "Descrição", "Situação"])
        for c in ws[ws.max_row]:
            c.font = fonte_header
            c.alignment = alin_centro
            c.border = borda

        for p in produtos_importados:
            nome = p.get("nome")
            preco_texto = ""
            preco_val = p.get("preco")
            nome_norm = normalizar_texto(nome)

            found = ""
            for k, info in primeira_ocorrencia_info.items():
                if k[0] == nome_norm:
                    found = info.get("preco_texto", "")
                    break

            preco_texto = found if found is not None else ""

            situacao = "✅ Completo"
            faltando = []
            if not p.get("categoria"):
                faltando.append("categoria")
            if not p.get("marca"):
                faltando.append("marca")
            if not p.get("descricao"):
                faltando.append("descrição")
            if faltando:
                situacao = f"⚠️ Adicione {', '.join(faltando)} depois."

            ws.append([
                nome,
                preco_texto or "(não disponível)",
                preco_val if preco_val is not None else "(não convertido)",
                p.get("categoria") or "(vazio)",
                p.get("marca") or "(vazio)",
                p.get("descricao") or "(sem descrição)",
                situacao
            ])

            for c in ws[ws.max_row]:
                c.font = fonte_normal
                c.alignment = alin_esquerda
                c.border = borda

        # ===== Produtos Duplicados (aba separada) =====
        if duplicados_internos:
            ws_dup = wb.create_sheet("Produtos Duplicados")

            ws_dup.append([
                "Linha (duplicada)",
                "Nome (duplicada)",
                "Marca (duplicada)",
                "Primeira ocorrência (linha)",
                "Nome (primeira)",
                "Marca (primeira)",
                "Categoria (primeira)",
                "Preço (texto primeira)",
                "Preço (valor primeira)",
                "Descrição (primeira)"
            ])

            for c in ws_dup[ws_dup.max_row]:
                c.font = fonte_header
                c.alignment = alin_centro
                c.border = borda

            for d in duplicados_internos:
                linha_dup = d.get("linha_duplicada")
                nome_dup = d.get("nome")
                marca_dup = d.get("marca")
                primeira_linha = d.get("primeira_ocorrencia")

                nome_norm = normalizar_texto(nome_dup)
                marca_norm = normalizar_texto("" if marca_dup == "(vazio)" else marca_dup)

                chave_primeira = None
                if (nome_norm, marca_norm) in primeira_ocorrencia_info:
                    chave_primeira = (nome_norm, marca_norm)
                else:
                    for k in primeira_ocorrencia_info.keys():
                        if k[0] == nome_norm:
                            chave_primeira = k
                            break

                dados_primeira = primeira_ocorrencia_info.get(chave_primeira, {})

                ws_dup.append([
                    linha_dup,
                    nome_dup,
                    marca_dup,
                    primeira_linha or "",
                    dados_primeira.get("nome", ""),
                    dados_primeira.get("marca", ""),
                    dados_primeira.get("categoria", ""),
                    dados_primeira.get("preco_texto", ""),
                    dados_primeira.get("preco", ""),
                    dados_primeira.get("descricao", "")
                ])

                for c in ws_dup[ws_dup.max_row]:
                    c.font = fonte_normal
                    c.alignment = alin_esquerda
                    c.border = borda

            ws_dup.insert_rows(1)
            ws_dup["A1"].value = "Resumo e instruções rápidas sobre duplicados"
            ws_dup["A1"].font = Font(bold=True, size=12)
            ws_dup.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

            ws_dup.insert_rows(2)
            ws_dup["A2"].value = (
                "Para cada linha duplicada, verifique Marca e Preço original. "
                "Preencha marca vazia com 'SEM MARCA' ou a real. "
                "Se o preço veio sem vírgula (ex.: 299), confirme se é 299 ou 2,99."
            )
            ws_dup["A2"].alignment = alin_esquerda

        # ===== Aba de Dicas =====
        ws_dicas = wb.create_sheet("Dicas e Uso")
        ws_dicas.append(["Guia Rápido — Como usar este relatório"])
        ws_dicas["A1"].font = Font(bold=True, size=14)
        ws_dicas.append([])

        guia = [
            "1) Leia o Resumo primeiro: mostra quantos itens chegaram e quantos têm problemas.",
            "2) Abra 'Produtos com Erro' para ver linha por linha (texto simples + o que fazer).",
            "3) Em 'Produtos importados' veja os itens inseridos e confirme preços.",
            "4) Em 'Produtos Duplicados' compare linhas disputadas.",
            "5) Preços: o relatório mantém o valor textual original.",
            "6) Se um preço vier '299', confirme se é 299 ou 2,99.",
            "7) Se houver marca vazia, preencha com 'SEM MARCA' ou a correta."
        ]
        for l in guia:
            ws_dicas.append([l])

        ws_dicas.append([])
        ws_dicas.append(["Check-list rápido antes de importar"])

        checklist = [
            "✓ Cada produto tem NOME.",
            "✓ Preço preenchido.",
            "✓ Marca preenchida quando possível.",
            "✓ Arquivo salvo em .xlsx ou .csv (UTF-8)."
        ]
        for c in checklist:
            ws_dicas.append([c])

        for sheet in [ws]:
            for i, col in enumerate(sheet.columns, start=1):
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                sheet.column_dimensions[get_column_letter(i)].width = min(max_len + 4, 120)

        if 'ws_dicas' in locals():
            for i, col in enumerate(ws_dicas.columns, start=1):
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws_dicas.column_dimensions[get_column_letter(i)].width = min(max_len + 4, 120)

        if 'ws_dup' in locals():
            for i, col in enumerate(ws_dup.columns, start=1):
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws_dup.column_dimensions[get_column_letter(i)].width = min(max_len + 4, 120)

        output = os.path.join(
            "static/relatorios",
            f"Relatorio_Importacao_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        os.makedirs(os.path.dirname(output), exist_ok=True)
        wb.save(output)

        return send_file(
            output,
            as_attachment=True,
            download_name="Relatorio_Importacao.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        traceback.print_exc()
        print("ERRO NA IMPORTAÇÃO:", e)
        print(traceback.format_exc())
        return jsonify({"sucesso": False, "erro": str(e)}), 500

    finally:
        if wb_raw is not None:
            try:
                wb_raw.close()
            except Exception:
                pass
        if caminho_temp and os.path.exists(caminho_temp):
            os.remove(caminho_temp)




# -----------------------------
# Deletar múltiplos produtos selecionados
# -----------------------------
@comerciante_bp.route("/api/produtos/deletar_selecionados", methods=["POST"])
@login_required
def deletar_selecionados():
    import time
    import traceback

    BLOCO = 500
    MAX_RETRIES = 3

    try:
        data = request.get_json(force=True) or {}
        ids = data.get("ids", [])
        if not isinstance(ids, list) or not ids:
            return jsonify({"sucesso": False, "erro": "Envie uma lista válida de IDs."}), 400

        ids_validos = list({str(i).strip() for i in ids if isinstance(i, str) and i.strip()})
        if not ids_validos:
            return jsonify({"sucesso": False, "erro": "Nenhum ID válido fornecido."}), 400

        comerciante_id = session.get("comerciante_id")
        if not comerciante_id:
            return jsonify({"sucesso": False, "erro": "Sessão expirada."}), 401

        # Buscar produtos existentes com imagens
        existentes = supabase.table("produtos") \
            .select("id, imagem") \
            .in_("id", ids_validos) \
            .eq("comerciante_id", comerciante_id) \
            .execute()

        produtos_existentes = existentes.data or []
        if not produtos_existentes:
            return jsonify({"sucesso": False, "erro": "Nenhum produto encontrado."}), 404

        total_excluidos = 0
        lotes_falhos = []

        for i in range(0, len(produtos_existentes), BLOCO):
            lote = produtos_existentes[i:i+BLOCO]

            # 🔹 Excluir imagens do Supabase
            for p in lote:
                imagem_url = p.get("imagem")
                if imagem_url:
                    nome_arquivo = imagem_url.split("/")[-1]
                    try:
                        supabase.storage.from_("produtos").remove([nome_arquivo])
                        print(f"Imagem {nome_arquivo} removida com sucesso.")
                    except Exception as e_img:
                        print(f"⚠️ Erro ao remover imagem {nome_arquivo}: {e_img}")

            # 🔹 Excluir produtos do banco
            for tentativa in range(1, MAX_RETRIES + 1):
                try:
                    supabase.table("produtos") \
                        .delete() \
                        .in_("id", [p["id"] for p in lote]) \
                        .eq("comerciante_id", comerciante_id) \
                        .execute()
                    total_excluidos += len(lote)
                    print(f"Lote {i//BLOCO + 1} excluído com sucesso.")
                    break
                except Exception as e:
                    print(f"⚠️ Erro no lote {i//BLOCO + 1} (tentativa {tentativa}): {e}")
                    time.sleep(0.2 * tentativa)
            else:
                lotes_falhos.append([p["id"] for p in lote])
                print(f"❌ Lote {i//BLOCO + 1} falhou após {MAX_RETRIES} tentativas.")

            time.sleep(0.01)

        return jsonify({
            "sucesso": True,
            "mensagem": f"{total_excluidos} produto(s) excluído(s) com sucesso.",
            "lotes_falhos": lotes_falhos
        }), 200

    except Exception as e:
        print("❌ Erro ao deletar produtos:", str(e))
        traceback.print_exc()
        return jsonify({"sucesso": False, "erro": "Erro interno ao processar a exclusão."}), 500



@comerciante_bp.route("/api/produtos/atualizar", methods=["POST"])
@login_required
def atualizar_produtos():
    import pandas as pd
    import chardet, traceback, os, unicodedata, xml.etree.ElementTree as ET, numpy as np
    from werkzeug.utils import secure_filename
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import jsonify, request, session, send_file
    import tempfile, shutil
    import io

    caminho_temp = None
    try:
        # ===================== Valida upload =====================
        if "arquivo" not in request.files:
            return jsonify({"sucesso": False, "erro": "Nenhum arquivo enviado."}), 400

        arquivo = request.files["arquivo"]
        if arquivo.filename == "":
            return jsonify({"sucesso": False, "erro": "Nenhum arquivo selecionado."}), 400

        filename = secure_filename(arquivo.filename)
        extensao = filename.lower().split(".")[-1]
        caminho_temp = os.path.join("static/uploads/temp", filename)
        os.makedirs(os.path.dirname(caminho_temp), exist_ok=True)
        arquivo.save(caminho_temp)

        # ===================== Leitura do arquivo =====================
        if extensao in ["xlsx", "xls"]:
            df = pd.read_excel(caminho_temp)

        elif extensao == "csv":
            with open(caminho_temp, "rb") as f:
                result = chardet.detect(f.read(10000))
            encoding = result["encoding"] or "utf-8"
            tentativas = [";", ",", "\t", "|"]

            for sep in tentativas:
                try:
                    df = pd.read_csv(caminho_temp, encoding=encoding, sep=sep)
                    if len(df.columns) > 1 or len(df) > 0:
                        break
                except Exception:
                    continue

        elif extensao == "json":
            df = pd.read_json(caminho_temp)

        elif extensao == "xml":
            tree = ET.parse(caminho_temp)
            root = tree.getroot()
            data = [{child.tag: child.text for child in item} for item in root]
            df = pd.DataFrame(data)

        elif extensao == "txt":
            with open(caminho_temp, "rb") as f:
                result = chardet.detect(f.read(10000))
            encoding = result["encoding"] or "utf-8"

            with open(caminho_temp, "r", encoding=encoding, errors="ignore") as f:
                primeira_linha = f.readline()

            if any(x in primeira_linha for x in [";", ",", "\t"]):
                tentativas = [";", ",", "\t", "|"]
                for sep in tentativas:
                    try:
                        df = pd.read_csv(caminho_temp, encoding=encoding, sep=sep)
                        if len(df.columns) > 1 or len(df) > 0:
                            break
                    except Exception:
                        continue
            else:
                linhas = [
                    l.strip()
                    for l in open(caminho_temp, encoding=encoding, errors="ignore").readlines()
                    if l.strip()
                ]
                df = pd.DataFrame({"linha": linhas})

        elif extensao == "ods":
            df = pd.read_excel(caminho_temp, engine="odf")

        else:
            return jsonify({"sucesso": False, "erro": f"Formato de arquivo não suportado: .{extensao}"}), 400

        if df.empty:
            return jsonify({"sucesso": False, "erro": "Arquivo vazio."}), 400

        # ===================== Normalização =====================
        def normalizar_coluna(c):
            c = str(c).strip().lower()
            c = unicodedata.normalize("NFKD", c).encode("ASCII", "ignore").decode("utf-8")
            return c

        df.columns = [normalizar_coluna(c) for c in df.columns]
        df = df.replace({np.nan: None})

        comerciante_id = session.get("comerciante_id")
        imagem_padrao = "/static/img/sem-imagem.png"

        produtos_atualizados = []
        erros = []

        # ===================== Produtos existentes =====================
        resp = supabase.table("produtos").select("*").eq("comerciante_id", comerciante_id).execute()
        produtos_existentes = {p["nome"].strip().lower(): p for p in (resp.data or [])}

        # ===================== Função imagem =====================
        def processar_imagem(valor):
            if not valor or str(valor).strip().lower() in ["nan", "none"]:
                return imagem_padrao

            valor = str(valor).strip()

            if valor.startswith("http://") or valor.startswith("https://"):
                return valor

            local = os.path.join("static/uploads", valor)
            if os.path.exists(local):
                return "/" + local.replace(os.sep, "/")

            return imagem_padrao

        # ===================== Processamento =====================
        for idx, row in df.iterrows():
            linha = idx + 2
            try:
                nome = str(row.get("nome", "")).strip()

                if not nome:
                    erros.append({
                        "linha": linha,
                        "erro": "Nome ausente. Produto não identificado.",
                        "instrucao": "Preencha o campo 'nome'."
                    })
                    continue

                nome_key = nome.lower()

                if nome_key not in produtos_existentes:
                    erros.append({
                        "linha": linha,
                        "erro": f"Produto '{nome}' não encontrado no sistema.",
                        "instrucao": "Verifique se o nome está correto ou se já existe no sistema."
                    })
                    continue

                produto = produtos_existentes[nome_key]

                preco_raw = str(row.get("preco", "")).replace("R$", "").replace(",", ".").strip()

                try:
                    preco = float(preco_raw) if preco_raw not in ["", "nan", "None"] else produto.get("preco", 0.0)
                except ValueError:
                    preco = produto.get("preco", 0.0)
                    erros.append({
                        "linha": linha,
                        "erro": f"Preço inválido '{row.get('preco', '')}'. Mantido valor anterior.",
                        "instrucao": "Corrija o formato do preço (ex: 10.99)."
                    })

                dados_atualizados = {
                    "preco": preco,
                    "descricao": str(row.get("descricao", produto.get("descricao", "")) or "").strip(),
                    "categoria": str(row.get("categoria", produto.get("categoria", "")) or "").strip(),
                    "marca": str(row.get("marca", produto.get("marca", "")) or "").strip(),
                    "imagem": processar_imagem(row.get("imagem", produto.get("imagem", imagem_padrao))),
                    "atualizado_em": datetime.utcnow().isoformat()
                }

                supabase.table("produtos").update(dados_atualizados).eq("id", produto["id"]).execute()

                produtos_atualizados.append({
                    "nome": nome,
                    "preco_antigo": produto.get("preco", ""),
                    "preco_novo": preco,
                    "categoria": dados_atualizados["categoria"],
                    "marca": dados_atualizados["marca"]
                })

            except Exception as e:
                erros.append({
                    "linha": linha,
                    "erro": f"Erro inesperado: {str(e)}",
                    "instrucao": "Verifique os dados e tente novamente."
                })

        # =====================================================================
        #   RELATÓRIO EXCEL PADRONIZADO (USANDO O MODELO DE IMPORTAÇÃO)
        # =====================================================================

        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Atualização"

        cor_titulo = "2F4865"
        cor_secao = "D9E1F2"
        cor_texto = "1F1F1F"
        cor_erro = "FFDDDD"
        cor_erro_header = "C00000"
        cor_aviso = "FFF6CC"
        cor_sucesso = "E8F6EF"
        cor_branco = "FFFFFF"

        fonte_titulo = Font(bold=True, size=16, color="FFFFFF")
        fonte_header = Font(bold=True, color=cor_texto)
        fonte_normal = Font(size=11, color=cor_texto)
        fonte_secao = Font(bold=True, size=12, color="000000")
        fonte_dica = Font(italic=True, size=11, color="000000")
        alin_esquerda = Alignment(horizontal="left", vertical="top", wrap_text=True)
        alin_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
        borda = Border(
            left=Side(style="thin", color="999999"),
            right=Side(style="thin", color="999999"),
            top=Side(style="thin", color="999999"),
            bottom=Side(style="thin", color="999999")
        )

        # limpa sheet inicial (garantir vazia)
        try:
            ws.delete_rows(1, ws.max_row)
        except Exception:
            pass

        # título semelhante ao modelo de importação, porém adaptado
        ws.merge_cells("A1:G1")
        ws["A1"].value = "Relatório Analítico de Atualização de Produtos - 🔁 Preço Mais Barato"
        ws["A1"].font = fonte_titulo
        ws["A1"].alignment = alin_centro
        ws["A1"].fill = PatternFill(start_color=cor_titulo, end_color=cor_titulo, fill_type="solid")
        ws.row_dimensions[1].height = 26

        # === CABEÇALHO VERTICAL ===
        cab = {
            "Data do processamento": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            "Tipo de operação": "ATUALIZAÇÃO DE PRODUTOS",
            "Sistema": "Preço Mais Barato - Núcleo de Inteligência",
            "Versão do motor": "2.3.1",
            "Relatório": "Relatório Analítico de Atualização de Produtos"
        }
        for k, v in cab.items():
            ws.append([k, v])
            for cell in ws[ws.max_row]:
                cell.font = fonte_normal
                cell.alignment = alin_esquerda
        ws.append([])

        # ===== Resumo =====
        ws.append(["Resumo da Atualização"])
        ws["A" + str(ws.max_row)].font = fonte_secao
        ws["A" + str(ws.max_row)].fill = PatternFill(start_color=cor_secao, end_color=cor_secao, fill_type="solid")

        ws.append(["Total de linhas", "Atualizados", "Com Erro", "Com Aviso"])
        for c in ws[ws.max_row]:
            c.font = fonte_header
            c.alignment = alin_centro
            c.border = borda

        # contagem de avisos se houver campo 'tipo' == 'aviso' em erros; caso contrário 0
        avisos_count = len([e for e in erros if e.get("tipo") == "aviso"]) if erros else 0

        ws.append([
            len(df),
            len(produtos_atualizados),
            len([e for e in erros if e.get("tipo") == "erro"]) if erros else len(erros),
            avisos_count
        ])
        for c in ws[ws.max_row]:
            c.alignment = alin_centro
            c.border = borda

        ws.append([])

        # ===== Produtos com Erro / Avisos =====
        ws.append(["⚠️ Produtos com Erro / Avisos (Simples ➜ Solução)"])
        ws["A" + str(ws.max_row)].font = fonte_secao
        ws["A" + str(ws.max_row)].fill = PatternFill(start_color=cor_secao, end_color=cor_secao, fill_type="solid")

        if erros:
            ws.append(["Linha", "O que aconteceu (simples)", "O que fazer (simples)", "Análise técnica (código)"])
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = alin_centro
                cell.fill = PatternFill(start_color=cor_erro_header, end_color=cor_erro_header, fill_type="solid")
                cell.border = borda

            # agrupar erros por linha
            erros_por_linha = {}
            for e in erros:
                l = e.get("linha")
                msg = e.get("erro", "")
                if l not in erros_por_linha:
                    erros_por_linha[l] = []
                if msg not in erros_por_linha[l]:
                    erros_por_linha[l].append(msg)

            def mapear_instrucoes(joined_msg):
                lower_msg = joined_msg.lower()
                instrucoes = []
                instr_tecnica = []

                if "nome ausente" in lower_msg or "sem nome" in lower_msg:
                    instrucoes.append("Escreva o NOME do produto nessa linha.")
                    instr_tecnica.append(("E-01", "Nome ausente: campo obrigatório."))

                if "preço ausente" in lower_msg or "preço inválido" in lower_msg:
                    instrucoes.append("Preencha o PREÇO (ex.: 2,99 ou 299).")
                    instr_tecnica.append(("E-03", "Preço ausente/inválido: impede atualização."))

                if "não encontrado" in lower_msg or "produto" in lower_msg and "não encontrado" in lower_msg:
                    instrucoes.append("Verifique se o NOME está correto ou se o produto existe no sistema.")
                    instr_tecnica.append(("E-05", "Produto não localizado: não existe para atualizar."))

                if not instrucoes:
                    instrucoes.append("Corrija os campos indicados e tente novamente.")
                    instr_tecnica.append(("E-00", "Erro genérico: verificar manualmente."))

                return " ".join(instrucoes), " | ".join([f"{c}: {d}" for c, d in instr_tecnica])

            for linha_idx in sorted(erros_por_linha.keys()):
                mensagens = erros_por_linha[linha_idx]
                joined_msg = "; ".join(mensagens)
                instr_simples, instr_tecnica = mapear_instrucoes(joined_msg)

                ws.append([linha_idx, joined_msg, instr_simples, instr_tecnica])
                last_row = ws.max_row

                fill = PatternFill(start_color=cor_erro, end_color=cor_erro, fill_type="solid")
                if "marca ausente" in joined_msg.lower():
                    fill = PatternFill(start_color=cor_aviso, end_color=cor_aviso, fill_type="solid")

                for cell in ws[last_row]:
                    cell.alignment = alin_esquerda
                    cell.font = fonte_normal
                    cell.border = borda
                    cell.fill = fill

            ws.append([])
        else:
            ws.append(["✅ Nenhum erro encontrado."])
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True, color="000000")
                cell.alignment = alin_centro

        ws.append([])

        # ===== Produtos Atualizados com Sucesso =====
        ws.append(["✅ Produtos Atualizados com Sucesso"])
        ws["A" + str(ws.max_row)].font = fonte_secao
        ws["A" + str(ws.max_row)].fill = PatternFill(start_color=cor_sucesso, end_color=cor_sucesso, fill_type="solid")

        ws.append(["Nome", "Preço Anterior", "Novo Preço", "Categoria", "Marca", "Situação"])
        for c in ws[ws.max_row]:
            c.font = fonte_header
            c.alignment = alin_centro
            c.border = borda

        for p in produtos_atualizados:
            ws.append([
                p.get("nome", ""),
                p.get("preco_antigo", ""),
                p.get("preco_novo", ""),
                p.get("categoria", "(vazio)"),
                p.get("marca", "(vazio)"),
                "Atualizado"
            ])
            for c in ws[ws.max_row]:
                c.font = fonte_normal
                c.alignment = alin_esquerda
                c.border = borda

        # ===== Aba de Dicas =====
        ws_dicas = wb.create_sheet("Dicas e Uso")
        ws_dicas.append(["Guia Rápido — Como usar este relatório"])
        ws_dicas["A1"].font = Font(bold=True, size=14)
        ws_dicas.append([])
        guia = [
            "1) Leia o Resumo primeiro: mostra quantos itens chegaram e quantos têm problemas.",
            "2) Abra 'Produtos com Erro' para ver linha por linha (texto simples + o que fazer).",
            "3) Em 'Produtos Atualizados' veja os itens atualizados e confirme preços.",
            "4) Se houve 'Produto não encontrado', corrija o nome no sistema e tente novamente.",
            "5) Preços: o relatório mantém o valor textual original quando disponível.",
            "6) Se um preço vier '299', confirme se é 299 ou 2,99.",
            "7) Se houver marca vazia, preencha com 'SEM MARCA' ou a correta."
        ]
        for l in guia:
            ws_dicas.append([l])

        ws_dicas.append([])
        ws_dicas.append(["Check-list rápido antes de atualizar"])
        checklist = [
            "✓ Cada produto tem NOME.",
            "✓ Preço preenchido quando for alterar.",
            "✓ Marca preenchida quando possível.",
            "✓ Arquivo salvo em .xlsx ou .csv (UTF-8)."
        ]
        for c in checklist:
            ws_dicas.append([c])

        # ajustar largura das colunas
        for sheet in [ws, ws_dicas]:
            for i, col in enumerate(sheet.columns, start=1):
                try:
                    max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                except Exception:
                    max_len = 0
                sheet.column_dimensions[get_column_letter(i)].width = min(max_len + 4, 120)

        # salvar em memória e enviar (BytesIO)
        output = io.BytesIO()
        wb.save(output)
        wb.close()
        output.seek(0)

        nome_arquivo = f"Relatorio_Atualizacao_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        # não imprimir stacktrace aqui para não sujar a resposta binária
        return jsonify({"sucesso": False, "erro": "Erro interno ao processar o arquivo."}), 500

    finally:
        if caminho_temp and os.path.exists(caminho_temp):
            try:
                os.remove(caminho_temp)
            except:
                pass


            
@comerciante_bp.route("/api/upload", methods=["POST"])
@login_required
def api_upload_imagem():
    import traceback

    print("\n========================")
    print("📥 DEBUG UPLOAD INICIADO")
    print("========================")

    try:
        # Verifica se veio arquivo
        print("🔍 request.files keys:", list(request.files.keys()))
        print("🔍 request.form keys:", list(request.form.keys()))

        file = request.files.get("file")  # o nome deve bater com o do FormData
        if not file:
            return jsonify({"erro": "Nenhum arquivo enviado"}), 400

        print("📌 Arquivo encontrado no campo:", "file")
        print("📄 Nome original:", file.filename)
        print("📦 Tamanho do arquivo recebido:", len(file.read()), "bytes")
        file.seek(0)  # volta para o início do arquivo após ler

        # Gera nome único
        nome_arquivo = f"{uuid.uuid4()}.jpg"
        print("📝 Nome gerado:", nome_arquivo)

        # Conteúdo binário
        conteudo = file.read()
        file.seek(0)

        # Envia para Supabase com MIME correto
        print("☁️ Enviando ao Supabase...")
        resultado = supabase.storage.from_("produtos").upload(
            nome_arquivo,
            conteudo,
            {"content-type": file.content_type}  # garante o tipo correto
        )

        if "error" in str(resultado).lower():
            print("❌ ERRO no Supabase:", resultado)
            return jsonify({"erro": "Falha no upload", "detalhes": str(resultado)}), 500

        # Pega URL pública
        url_publica = supabase.storage.from_("produtos").get_public_url(nome_arquivo)
        print("✅ Upload concluído. URL pública:", url_publica)

        return jsonify({
            "sucesso": True,
            "url": url_publica,
            "arquivo": nome_arquivo
        })

    except Exception as e:
        print("🔥 EXCEÇÃO:", str(e))
        traceback.print_exc()
        return jsonify({"erro": str(e)}), 500
    
    
@comerciante_bp.route("/api/produtos/upload_imagem", methods=["POST"])
@login_required
def upload_imagem_produto():
    import uuid, os, traceback
    from flask import request, jsonify

    try:
        # ===========================================================
        # 1) PEGAR CAMPO "imagem" DO FORM-DATA
        # ===========================================================
        imagem_valor = request.form.get("imagem", "").strip()

        # ===========================================================
        # 2) SE FOR URL — RETORNA DIRETO
        # ===========================================================
        if imagem_valor.startswith("http://") or imagem_valor.startswith("https://"):
            return jsonify({
                "sucesso": True,
                "origem": "url",
                "url_final": imagem_valor
            })

        # ===========================================================
        # 3) SE FOR NOME DE ARQUIVO LOCAL — JÁ EXISTE NO SUPABASE
        # ===========================================================
        if imagem_valor.endswith(".jpg") or imagem_valor.endswith(".png") or imagem_valor.endswith(".jpeg"):
            url_publica = supabase.storage.from_("produtos").get_public_url(imagem_valor)
            return jsonify({
                "sucesso": True,
                "origem": "nome_existente",
                "url_final": url_publica
            })

        # ===========================================================
        # 4) SE FOR UPLOAD REAL - “file”
        # ===========================================================
        arquivo = request.files.get("file")

        if not arquivo:
            return jsonify({
                "sucesso": False,
                "erro": "Nenhuma imagem enviada (campo 'file')."
            }), 400

        # Lê o conteúdo binário
        conteudo = arquivo.read()

        # Gera nome único no Supabase
        ext = os.path.splitext(arquivo.filename)[1].lower()
        if ext not in [".jpg", ".jpeg", ".png", ".webp"]:
            ext = ".jpg"

        nome_arquivo = f"{uuid.uuid4()}{ext}"

        # Upload com MIME-type correto
        resultado = supabase.storage.from_("produtos").upload(
            nome_arquivo,
            conteudo,
            {"content-type": arquivo.content_type}
        )

        if "error" in str(resultado).lower():
            return jsonify({"sucesso": False, "erro": "Falha ao enviar imagem ao Storage."}), 500

        # Pega URL pública
        url_publica = supabase.storage.from_("produtos").get_public_url(nome_arquivo)

        return jsonify({
            "sucesso": True,
            "origem": "upload",
            "url_final": url_publica,
            "arquivo": nome_arquivo
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({
            "sucesso": False,
            "erro": str(e)
        }), 500


def upload_foto_comerciante(file):
    if not file or file.filename == "":
        return None

    # Gera nome único
    ext = file.filename.rsplit(".", 1)[-1].lower()
    filename = f"{uuid.uuid4()}.{ext}"

    caminho_supabase = f"perfil/{filename}"  # pasta perfil dentro do bucket comerciante

    # Lê bytes
    file_bytes = file.read()

    # Upload para o bucket
    res = supabase.storage.from_("comerciante").upload(
        caminho_supabase,
        file_bytes,
        {"content-type": file.content_type}
    )

    if res is None:
        return None

    # Gera URL pública
    public_url = supabase.storage.from_("comerciante").get_public_url(caminho_supabase)

    return public_url
